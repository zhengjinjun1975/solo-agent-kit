# -*- coding: utf-8 -*-
"""excel_report.py — Excel 交付报告（openpyxl，本机已装）。

FDE 现场分析/清洗结果生成 Excel 交付客户。
三个能力：清洗报告 / 分析报告 / 本体导出。

依赖 openpyxl；不可用时明确报错（可降级）。
"""
from __future__ import annotations

import os

try:
    from openpyxl import Workbook
    _XLSX = True
except ImportError:
    _XLSX = False


def _out_dir() -> str:
    d = os.environ.get("SOLO_REPORTS", os.path.expanduser("~/.solo/reports"))
    os.makedirs(d, exist_ok=True)
    return d


def _require_xlsx():
    if not _XLSX:
        raise RuntimeError("openpyxl 未安装（可选依赖），无法生成 Excel。请 pip install openpyxl")


def clean_report(data: list, path: str = None) -> dict:
    """清洗前后对比报告。

    data: [{原值, 清洗后, 是否异常, 原因, ...}, ...]
    path: 输出路径，None 用 ~/.solo/reports/clean_report.xlsx
    """
    _require_xlsx()
    wb = Workbook()
    ws = wb.active
    ws.title = "清洗报告"
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row.get(h, "") for h in headers])
    out = path or os.path.join(_out_dir(), "clean_report.xlsx")
    wb.save(out)
    return {"ok": True, "path": out.replace("\\", "/"), "rows": len(data)}


def analysis_report(stats: dict, path: str = None) -> dict:
    """统计指标报告。

    stats: {字段: {mean, median, std, min, max, anomalies}, ...}（factory.stats.describe 输出）
    path: 输出路径
    """
    _require_xlsx()
    wb = Workbook()
    ws = wb.active
    ws.title = "分析报告"
    headers = ["字段", "均值", "中位数", "标准差", "最小值", "最大值", "异常数"]
    ws.append(headers)
    for field, s in (stats or {}).items():
        ws.append([field, s.get("mean"), s.get("median"), s.get("std"),
                   s.get("min"), s.get("max"), s.get("anomalies", 0)])
    out = path or os.path.join(_out_dir(), "analysis_report.xlsx")
    wb.save(out)
    return {"ok": True, "path": out.replace("\\", "/"),
            "rows": len(stats or {})}


def ontology_report(entities: list, relations: list, path: str = None) -> dict:
    """本体导出报告：实体表 + 关系表 两个 sheet。

    entities: [{id, type, name, ...}, ...]
    relations: [{subject, predicate, object}, ...]
    """
    _require_xlsx()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "实体"
    if entities:
        h = list(entities[0].keys())
        ws1.append(h)
        for e in entities:
            ws1.append([e.get(k, "") for k in h])
    ws2 = wb.create_sheet("关系")
    if relations:
        h = list(relations[0].keys()) if relations else ["subject", "predicate", "object"]
        ws2.append(h)
        for r in relations:
            ws2.append([r.get(k, "") for k in h])
    out = path or os.path.join(_out_dir(), "ontology_report.xlsx")
    wb.save(out)
    return {"ok": True, "path": out.replace("\\", "/"),
            "entities": len(entities), "relations": len(relations)}


def acceptance_report(items: list, title: str = "验收清单", path: str = None) -> dict:
    """验收清单/签收单（survey 模块验收阶段导出）。

    items: [{aid, rid, title, clause, result, evidence}, ...]
    title: sheet 名（默认"验收清单"）
    自动统计通过/未通过/待验收 行数写入表尾。
    """
    _require_xlsx()
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "验收清单"
    headers = ["验收编号", "需求编号", "需求标题", "验收条款", "结果", "证据"]
    ws.append(headers)
    for it in items or []:
        ws.append([it.get("aid"), it.get("rid"), it.get("title"),
                   it.get("clause"), it.get("result"), it.get("evidence")])
    # 表尾统计
    cnt = {}
    for it in items or []:
        cnt[it.get("result")] = cnt.get(it.get("result"), 0) + 1
    ws.append([])
    ws.append(["统计", f"共 {len(items or [])} 条",
               f"通过 {cnt.get('通过', 0)} / 未通过 {cnt.get('未通过', 0)}"
               f" / 待验收 {cnt.get('待验收', 0)}"])
    out = path or os.path.join(_out_dir(), "acceptance_report.xlsx")
    wb.save(out)
    return {"ok": True, "path": out.replace("\\", "/"), "rows": len(items or [])}


def quote_report(quote: dict, path: str = None) -> dict:
    """报价单导出（quote 模块商务报价用）。明细 + 汇总 + 条款 三个 sheet。

    quote: build_quote 输出（含 lines/totals/terms/effort）。
    path: 输出路径，None 用 ~/.solo/reports/quote_<project>.xlsx
    """
    _require_xlsx()
    wb = Workbook()
    # 明细
    ws = wb.active
    ws.title = "报价明细"
    ws.append(["项目", "范围"])
    ws.append([quote.get("project"), quote.get("scope")])
    ws.append([])
    headers = ["报价项", "人天", "单价", "金额"]
    ws.append(headers)
    for ln in quote.get("lines", []):
        ws.append([ln.get("item"), ln.get("days"), ln.get("unit_price"), ln.get("amount")])
    ws.append([])
    ws.append(["小计", "", "", quote["totals"]["subtotal"]])
    ws.append(["税费", "", "", quote["totals"]["tax"]])
    ws.append(["含税总额", "", "", quote["totals"]["total"]])
    # 汇总
    ws2 = wb.create_sheet("汇总")
    ws2.append(["项目", quote.get("project")])
    ws2.append(["范围", quote.get("scope")])
    ws2.append(["总人天", quote["totals"]["days"]])
    ws2.append(["质量分", quote["effort"]["quality"]])
    ws2.append(["质量档位", quote["effort"]["quality_level"]])
    ws2.append(["复杂度", quote["effort"]["complexity"]])
    ws2.append(["报价时间", quote.get("quoted_at")])
    # 条款
    ws3 = wb.create_sheet("条款")
    ws3.append(["条款", "内容"])
    for k, v in quote.get("terms", {}).items():
        ws3.append([k, v])
    proj = str(quote.get("project", "quote"))
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in proj)[:30] or "quote"
    out = path or os.path.join(_out_dir(), f"quote_{safe}.xlsx")
    wb.save(out)
    return {"ok": True, "path": out.replace("\\", "/"),
            "total": quote["totals"]["total"], "lines": len(quote.get("lines", []))}
